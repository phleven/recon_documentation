import json
import re
from pathlib import Path
from collections import defaultdict, Counter
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

base = Path('.')
common = json.loads((base / 'common-definitions.schema.json').read_text(encoding='utf-8'))
columns = json.loads((base / 'benefitsiq_dev_schema_columns.json').read_text(encoding='utf-8'))

domain_files = {
    'person': 'person.schema.json',
    'employment': 'employment.schema.json',
    'relationship': 'relationship.schema.json',
}


def split_tokens(text):
    text = text.replace('[]', ' ')
    text = re.sub(r'([a-z0-9])([A-Z])', r'\1 \2', text)
    text = re.sub(r'[^A-Za-z0-9]+', ' ', text)
    parts = [p.lower() for p in text.split() if p]
    out = []
    for p in parts:
        out.extend([q for q in p.split('_') if q])
    return out


def norm(text):
    return ''.join(split_tokens(text))


REF_CACHE = {
    'common-definitions.schema.json': common,
}


def resolve_ref(ref):
    if '#/' in ref:
        file_part, pointer = ref.split('#/', 1)
    else:
        file_part, pointer = ref, ''
    file_part = file_part or 'common-definitions.schema.json'
    if file_part not in REF_CACHE:
        REF_CACHE[file_part] = json.loads((base / file_part).read_text(encoding='utf-8'))
    node = REF_CACHE[file_part]
    if pointer:
        for token in pointer.split('/'):
            node = node[token]
    return node


def node_type(node):
    if 'type' in node:
        t = node['type']
        if isinstance(t, list):
            return '|'.join(str(x) for x in t)
        return str(t)
    if 'const' in node:
        return 'const'
    if 'enum' in node:
        return 'enum'
    if 'anyOf' in node:
        return 'anyOf'
    return ''


def collect_fields(root_name, root):
    rows = []

    def walk(node, path, required_set=None, seen_refs=None):
        if seen_refs is None:
            seen_refs = set()

        if '$ref' in node:
            ref = node['$ref']
            if ref in seen_refs:
                return
            walk(resolve_ref(ref), path, required_set, seen_refs | {ref})
            return

        if 'anyOf' in node and not node.get('properties'):
            rows.append({
                'domain': root_name,
                'json_path': '.'.join(path) if path else root_name,
                'json_field': path[-1] if path else root_name,
                'required': 'Y' if required_set and path and path[-1] in required_set else 'N',
                'json_type': node_type(node),
                'ref': '',
            })
            return

        t = node.get('type')
        if t == 'object' or 'properties' in node:
            props = node.get('properties', {})
            req = set(node.get('required', []))
            for key, child in props.items():
                walk(child, path + [key], req, seen_refs)
            return

        if t == 'array' or ('items' in node and t != 'object'):
            items = node.get('items', {})
            arr_key = (path[-1] + '[]') if path else '[]'
            new_path = path[:-1] + [arr_key] if path else [arr_key]
            walk(items, new_path, required_set, seen_refs)
            return

        rows.append({
            'domain': root_name,
            'json_path': '.'.join(path) if path else root_name,
            'json_field': path[-1].replace('[]', '') if path else root_name,
            'required': 'Y' if required_set and path and path[-1].replace('[]', '') in required_set else 'N',
            'json_type': node_type(node),
            'ref': node.get('$ref', ''),
        })

    walk(root, [])
    uniq = []
    seen = set()
    for r in rows:
        k = (r['domain'], r['json_path'])
        if k not in seen:
            seen.add(k)
            uniq.append(r)
    return uniq


for c in columns:
    c['col_norm'] = norm(c['column'])
    c['col_tokens'] = set(split_tokens(c['column']))
    c['table_tokens'] = set(split_tokens(c['table']))

schema_preferences = {
    'person': ['person_management', 'member', 'relationship', 'employment', 'hbe', 'retirement'],
    'employment': ['employment', 'member', 'person_management', 'relationship', 'hbe', 'retirement'],
    'relationship': ['relationship', 'person_management', 'member', 'employment', 'hbe', 'retirement'],
}

manual_overrides = {
    ('person', 'productModule'): {
        'schema': 'hbe',
        'table': 'role_matrix',
        'column': 'product',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Constant value SmartDev mapped to product column for downstream lineage/reporting.',
    },
    ('person', 'sourceSystem'): {
        'schema': 'hbe',
        'table': 'app_change_trigger',
        'column': 'source',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Source lineage mapped to generic source column.',
    },
    ('person', 'medicareEligible'): {
        'schema': 'retirement',
        'table': 'scar_peehip',
        'column': 'medicare_eligibility_ind',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Semantic mapping to retirement Medicare eligibility indicator.',
    },
    ('person', 'medicareEligDate'): {
        'schema': 'retirement',
        'table': 'health_insurance_enrollment_plan',
        'column': 'part_d_eligiblity_date',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped Medicare eligibility date to available Medicare Part D eligibility date column.',
    },
    ('person', 'supervisorName'): {
        'confidence': 'unmapped',
        'match_basis': 'no-target-column',
        'notes': 'No supervisor name column found in exported metadata.',
    },
    ('person', 'annuitantClaimNumber'): {
        'schema': 'retirement',
        'table': 'interface_rds_discrepancy',
        'column': 'ben_health_insurance_claim_number',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped to available beneficiary health insurance claim number column.',
    },
    ('person', 'addresses[].streetAddress'): {
        'schema': 'person_management',
        'table': 'adr',
        'column': 'line1_address',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Primary address line mapping.',
    },
    ('person', 'addresses[].foreignAddressLine'): {
        'schema': 'person_management',
        'table': 'adr',
        'column': 'line4_address',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Foreign address freeform detail mapped to line4_address.',
    },
    ('employment', 'productModule'): {
        'schema': 'hbe',
        'table': 'role_matrix',
        'column': 'product',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Constant value SmartDev mapped to product column for downstream lineage/reporting.',
    },
    ('employment', 'sourceSystem'): {
        'schema': 'hbe',
        'table': 'app_change_trigger',
        'column': 'source',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Source lineage mapped to generic source column.',
    },
    ('employment', 'fullTimePercentage'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'fte_indicator',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Semantic mapping. Transform percentage to FTE indicator/business rule as needed.',
    },
    ('employment', 'workLocation'): {
        'schema': 'retirement',
        'table': 'agency_location',
        'column': 'location_code',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped work location to location_code with code translation as required.',
    },
    ('employment', 'region'): {
        'schema': 'hbe',
        'table': 'rating_region_county',
        'column': 'rating_region_cd',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Region code mapped to rating region code.',
    },
    ('employment', 'exempt'): {
        'schema': 'retirement',
        'table': 'state_tax_status',
        'column': 'state_tax_exempt_ind',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped employment exempt indicator to tax exemption indicator.',
    },
    ('employment', 'unionAffiliation'): {
        'schema': 'hbe',
        'table': '_temp_plan_variant_data',
        'column': 'union_member_type',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Union affiliation indicator mapped to union member type with boolean/code transform.',
    },
    ('employment', 'unionCode'): {
        'schema': 'hbe',
        'table': '_temp_plan_variant_data',
        'column': 'union_member_type',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Union code mapped to union member type domain with code-map transform.',
    },
    ('employment', 'jobClassification'): {
        'schema': 'retirement',
        'table': 'ecm_enrollment_document_data',
        'column': 'classification',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Best available classification column candidate.',
    },
    ('employment', 'payPlan'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'pay_group',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped pay plan to pay_group code.',
    },
    ('employment', 'payLevel'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'pay_grade',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped pay level to pay_grade.',
    },
    ('employment', 'specialEmployeeCode'): {
        'schema': 'retirement',
        'table': 'ztr_special_handling_reason',
        'column': 'special_handling_reason_code',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped to available special handling reason code domain.',
    },
    ('employment', 'expatriate'): {
        'schema': 'person_management',
        'table': 'person_addl_attr',
        'column': 'foreign_national_ind',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Expatriate mapped to foreign national indicator.',
    },
    ('employment', 'ficaCode'): {
        'schema': 'retirement',
        'table': 'tax_exemption',
        'column': 'tax_type_code',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped FICA code to tax type code domain with code translation.',
    },
    ('employment', 'tenure'): {
        'schema': 'employment',
        'table': 'work_units',
        'column': 'contract_period_value',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped tenure to contract period value with unit conversion business rules as needed.',
    },
    ('employment', 'pshbRateScheduleCode'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'rate_cat_code',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped PSHB rate schedule to rate category code.',
    },
    ('employment', 'exempt2'): {
        'schema': 'retirement',
        'table': 'payment_history_tax_detail',
        'column': 'fed_exemption_count',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped secondary exemption input to federal exemption count.',
    },
    ('employment', 'exempt3'): {
        'schema': 'retirement',
        'table': 'payment_history_tax_detail',
        'column': 'state_exemption_count',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped tertiary exemption input to state exemption count.',
    },
    ('employment', 'terminationNoaCode'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'termination_type',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Mapped termination NOA code to termination type code.',
    },
    ('employment', 'owcpTerminationIndicator'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'owcp_termination_rsn',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Derive indicator from OWCP termination reason presence/value.',
    },
    ('employment', 'owcpTransferInIndicator'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'transfer_flag',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Direct transfer indicator mapping.',
    },
    ('employment', 'owcpReinstatementIndicator'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'reinstatement_flag',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Direct reinstatement indicator mapping.',
    },
    ('employment', 'reEmployedAnnuitantIndicator'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 're_emplyd_ant',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Direct re-employed annuitant indicator mapping.',
    },
    ('employment', 'preTaxHealthBenefitIndicator'): {
        'schema': 'person_management',
        'table': 'person_organization_association',
        'column': 'pre_tax_ind',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Direct pre-tax indicator mapping.',
    },
    ('relationship', 'productModule'): {
        'schema': 'hbe',
        'table': 'role_matrix',
        'column': 'product',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Constant value SmartDev mapped to product column for downstream lineage/reporting.',
    },
    ('relationship', 'sourceSystem'): {
        'schema': 'hbe',
        'table': 'app_change_trigger',
        'column': 'source',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Source lineage mapped to generic source column.',
    },
    ('relationship', 'relatedPerson.nonEmployeeUsCitizen'): {
        'schema': 'person_management',
        'table': 'person_addl_attr',
        'column': 'nat_us_ctzn_code',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Map boolean/coded citizenship status using nat_us_ctzn_code translation.',
    },
    ('relationship', 'relatedPerson.addresses[].streetAddress'): {
        'schema': 'person_management',
        'table': 'adr',
        'column': 'line1_address',
        'confidence': 'high',
        'match_basis': 'manual-override',
        'notes': 'Related person address line mapping.',
    },
    ('relationship', 'relatedPerson.addresses[].foreignAddressLine'): {
        'schema': 'person_management',
        'table': 'adr',
        'column': 'line4_address',
        'confidence': 'medium',
        'match_basis': 'manual-override',
        'notes': 'Related person foreign address detail mapped to line4_address.',
    },
}


def map_field(row):
    field = row['json_field']
    path = row['json_path']
    domain = row['domain']

    field_tokens = set(split_tokens(field))
    path_tokens = set(split_tokens(path))
    field_norm = norm(field)

    scored = []
    pref_rank = {s: i for i, s in enumerate(schema_preferences.get(domain, []))}

    for c in columns:
        score = 0

        if c['col_norm'] == field_norm and field_norm:
            score += 120
        elif field_norm and (field_norm in c['col_norm'] or c['col_norm'] in field_norm):
            score += 55

        overlap = len(field_tokens & c['col_tokens'])
        if overlap:
            score += overlap * 18

        path_overlap = len(path_tokens & c['table_tokens'])
        if path_overlap:
            score += path_overlap * 10

        if field.lower().endswith('id') and c['column'].lower().endswith('_id'):
            score += 15

        if c['schema'] in pref_rank:
            score += max(0, (6 - pref_rank[c['schema']]) * 5)

        if score > 0:
            scored.append((score, c))

    if not scored:
        auto = {
            **row,
            'db_schema': '',
            'db_table': '',
            'db_column': '',
            'db_data_type': '',
            'match_score': 0,
            'confidence': 'unmapped',
            'match_basis': 'no-candidate',
            'suggested_alternatives': '',
            'notes': 'No close column candidate found from exported metadata.',
        }
    else:
        scored.sort(key=lambda x: x[0], reverse=True)
        top_score, top = scored[0]

        if top_score >= 130:
            conf = 'high'
        elif top_score >= 95:
            conf = 'medium'
        elif top_score >= 60:
            conf = 'low'
        else:
            conf = 'unmapped'

        if conf == 'unmapped':
            db_schema = db_table = db_column = db_data_type = ''
            basis = 'below-threshold'
            notes = 'Candidates exist but score below mapping threshold.'
        else:
            db_schema = top['schema']
            db_table = top['table']
            db_column = top['column']
            db_data_type = top['data_type']
            basis = 'name-token-score'
            notes = ''

        alts = []
        for s, cand in scored[1:4]:
            alts.append(f"{cand['schema']}.{cand['table']}.{cand['column']} ({s})")

        auto = {
            **row,
            'db_schema': db_schema,
            'db_table': db_table,
            'db_column': db_column,
            'db_data_type': db_data_type,
            'match_score': top_score,
            'confidence': conf,
            'match_basis': basis,
            'suggested_alternatives': '; '.join(alts),
            'notes': notes,
        }

    override = manual_overrides.get((row['domain'], row['json_path']))
    if not override:
        return auto

    out = dict(auto)
    if override.get('schema') and override.get('table') and override.get('column'):
        out['db_schema'] = override['schema']
        out['db_table'] = override['table']
        out['db_column'] = override['column']
        match = next(
            (c for c in columns if c['schema'] == override['schema'] and c['table'] == override['table'] and c['column'] == override['column']),
            None,
        )
        out['db_data_type'] = match['data_type'] if match else ''
        out['match_score'] = max(out['match_score'], 140)
    elif override.get('confidence') == 'unmapped':
        out['db_schema'] = ''
        out['db_table'] = ''
        out['db_column'] = ''
        out['db_data_type'] = ''
        out['match_score'] = 0

    out['confidence'] = override.get('confidence', out['confidence'])
    out['match_basis'] = override.get('match_basis', out['match_basis'])
    if override.get('notes'):
        out['notes'] = override['notes']
    return out


all_fields = []
for domain, fname in domain_files.items():
    root = json.loads((base / fname).read_text(encoding='utf-8'))
    all_fields.extend(collect_fields(domain, root))

mapped_rows = [map_field(r) for r in all_fields]

headers = [
    'Domain', 'JSON Path', 'JSON Field', 'Required', 'JSON Type',
    'Target Schema', 'Target Table', 'Target Column', 'Target Data Type',
    'Match Score', 'Confidence', 'Match Basis', 'Suggested Alternatives', 'Notes'
]

wb = Workbook()
ws_overview = wb.active
ws_overview.title = 'overview'

ws_overview.append(['Mapping Workbook', 'benefitsiq-dev JSON to DB table/column mapping'])
ws_overview.append(['Generated On', str(date.today())])
ws_overview.append(['Target Host', 'be-iq-npd-rds.cal882m02hcx.us-east-1.rds.amazonaws.com'])
ws_overview.append(['Metadata Source', 'benefitsiq_dev_tables.json + benefitsiq_dev_schema_columns.json'])
ws_overview.append([])
ws_overview.append(['Domain', 'Total Fields', 'Mapped', 'Unmapped', 'High', 'Medium', 'Low'])

by_domain = defaultdict(list)
for r in mapped_rows:
    by_domain[r['domain']].append(r)

for d in ['person', 'employment', 'relationship']:
    rows = by_domain[d]
    total = len(rows)
    mapped = sum(1 for x in rows if x['confidence'] != 'unmapped')
    unmapped = total - mapped
    cnt = Counter(x['confidence'] for x in rows)
    ws_overview.append([d, total, mapped, unmapped, cnt.get('high', 0), cnt.get('medium', 0), cnt.get('low', 0)])

for cell in ws_overview[1]:
    cell.font = Font(bold=True)
for cell in ws_overview[6]:
    cell.font = Font(bold=True)


def write_sheet(name, rows):
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E78')
        c.alignment = Alignment(vertical='center')

    for r in rows:
        ws.append([
            r['domain'], r['json_path'], r['json_field'], r['required'], r['json_type'],
            r['db_schema'], r['db_table'], r['db_column'], r['db_data_type'],
            r['match_score'], r['confidence'], r['match_basis'], r['suggested_alternatives'], r['notes']
        ])

    widths = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            value = '' if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(80, len(value) + 2))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'


write_sheet('field_map_all', mapped_rows)
write_sheet('person', by_domain['person'])
write_sheet('employment', by_domain['employment'])
write_sheet('relationship', by_domain['relationship'])
write_sheet('unmapped_only', [r for r in mapped_rows if r['confidence'] == 'unmapped'])

out = base / f"benefitsiq_dev_json_to_table_column_mapping_{date.today().strftime('%Y%m%d')}.xlsx"
wb.save(out)

print(out.name)
print(f"rows={len(mapped_rows)}")
print(f"unmapped={sum(1 for r in mapped_rows if r['confidence'] == 'unmapped')}")
